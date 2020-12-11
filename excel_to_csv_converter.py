#!/usr/bin/env python3
# excel_to_csv_converter.py - Converts excel excel files to csv files.
import csv
import openpyxl
import os


os.makedirs('csv_files', exist_ok=True)

for excel_file in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if not excel_file.endswith('.xlsx'):
        continue
    work_book = openpyxl.load_workbook(excel_file, data_only=True)
    for sheet_name in work_book.sheetnames:
        sheet = work_book[sheet_name]

        # Create the csv.writer object For the csv file.
        with open(f"./csv_files/{excel_file.strip('.xlsx')}_{sheet_name}.csv", 'w', newline='') as csv_out:
            csv_writer = csv.writer(csv_out)

            # Get cells from work sheet from each row and put them in a list.
            row_list = list()
            for row in range(1, sheet.max_row + 1):
                for cell in range(1, sheet.max_column + 1):
                    if sheet.cell(row, cell).value != None:
                        row_list.append(sheet.cell(row, cell).value)

                csv_writer.writerow(row_list)
                row_list.clear()




    # Loop through every sheet in the workbook.
