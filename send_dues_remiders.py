#! /usr/bin/env python3
# send_dues_remaiders.py - Sends emails based on payment status in spreadsheet.

from os import environ
import openpyxl
import smtplib
import sys

# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']
last_col = sheet.max_column
latest_month = sheet.cell(row=1, column=last_col).value

# Check each member's payment status.
client_email = dict()

for x in range(1, sheet.max_row):
    if sheet.cell(row=x, column=last_col).value is None:
        client_email[sheet.cell(row=x, column=1).value] = sheet.cell(row=x, column=2).value

# Log in to email account.
EMAIL_ADDRESS = environ.get("EMAIL_ADDRESS")
EMAIL_PASS = environ.get("EMAIL_PASS")
smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)
smtp_obj.ehlo()
smtp_obj.starttls()
smtp_obj.login(EMAIL_ADDRESS, EMAIL_PASS)

# Send out reminder emails.
for name, email in client_email.items():
    print(latest_month)
    print(name)
    print(latest_month)
    body = "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues for %s. Please make this " \
           "payment as soon as possible. Thank you!" % (latest_month, name, latest_month)

    print('Sending email to %s: ' % email, body)
    sendmail_status = smtp_obj.sendmail(EMAIL_ADDRESS, email, body)

    if sendmail_status != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmail_status))

smtp_obj.quit()
