import openpyxl


# Getting Sheets from a Workbook
wb = openpyxl.load_workbook('example.xlsx')
print(type(wb), '\n')
print(wb.sheetnames, '\n') # The workbook's sheet's names.
sheet = wb['Sheet3'] # Get a sheet form the workbook.
print(sheet, '\n')
print(sheet.title, '\n') # Get the sheet's title as a string.
another_sheet = wb.active # Get the active sheet.
print(another_sheet)


## Getting Cells from the Sheet
#wb = openpyxl.load_workbook('example.xlsx')
#sheet = wb['Sheet1'] # Get a sheet from a workbook.
#print(sheet['A1']) # Get a cell from the sheet.
#print(sheet['A1'].value) # Get the value from the cell.
#c = sheet['B1'] # Get another value from the cell.
#print(c.value)
#
#
## Get the row, column and value from the cell.
#print('Row %s, Column %s is %s' % (c.row, c.column, c.value))
#print('Cell %s is %s' % (c.coordinate, c.value))
#print(sheet['C1'].value, '\n')
#print(sheet.cell(row=1, column=2))
#print(sheet.cell(row=1, column=2).value, '\n')
#
#for i in range(1, 8, 2): # Go through every other row:
#    print(i, sheet.cell(row=i, column=2).value)
#
#wb = openpyxl.load_workbook('example.xlsx')
#sheet = wb['Sheet1']
#print(sheet.max_row) # Get the highest row number.
#print(sheet.max_column) # Get the highest column number.
#
#
## Converting Between Column Letters and Numbers
#from openpyxl.utils import get_column_letter, column_index_from_string
#
#print(get_column_letter(1))
#print(get_column_letter(2))
#print(get_column_letter(27))
#print(get_column_letter(900))
#wb = openpyxl.load_workbook('example.xlsx')
#sheet = wb['Sheet1']
#print(get_column_letter(sheet.max_column))
#print(column_index_from_string('A'))
#print(column_index_from_string('AA'))
#
#
## Getting Rows and Columns from the Sheets
#wb = openpyxl.load_workbook('example.xlsx')
#sheet = wb['Sheet1']
#print(tuple(sheet['A1':'C3']),'\n') # Get all cell from A1 to C3.
#for row_of_cell_objects in sheet['A1':'C3']:
#    print(row_of_cell_objects)
#    for cell_obj in row_of_cell_objects:
#        print(cell_obj.coordinate, cell_obj.value)
#    print('--- END OF ROW ---')
#
#
#wb = openpyxl.load_workbook('example.xlsx')
#sheet = wb.active
#print(list(sheet.columns)[1])
#print()
#for cell_obj in list(sheet.columns)[1]:
#    print(cell_obj.value)
#
#
