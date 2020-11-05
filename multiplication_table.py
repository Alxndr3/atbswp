#! /bin/env python3
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from sys import argv


if len(argv) < 2:
    print("Type one integer as argument:")

else:
    argv_value = int(argv[1])

    wb = openpyxl.Workbook()
    sheet = wb.active
    bold_12_font = Font(size=12, bold=True)

    for x in range(1, argv_value + 1):
        sheet['A' + str(x + 1)].font = bold_12_font
        sheet[get_column_letter(x + 1) + str(1)].font = bold_12_font
        sheet['A' + str(x + 1)] = x
        sheet[get_column_letter(x + 1) + str(1)] = x

        for y in range(1, argv_value + 1):
              sheet.cell(row=x+1, column=y+1).value = x * y


wb.save('multi_table.xlsx')




