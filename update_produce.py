import openpyxl
from time import time


start_time = time()

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices
PRICE_UPDATES = {
        'Garlic': 3.07,
        'Celery': 1.19,
        'Lemon': 1.27
        }

# Loop through the rows and update the prices.
# for row_num in range(2, sheet.max_row):
#     produce_name = sheet.cell(row=row_num, column=1).value
#     if produce_name in PRICE_UPDATES:
#         sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]
# wb.save('produce_sales_updated.xlsx')



# Loop through the rows and update the prices.
for row in range(2, sheet.max_row + 1):
    for k, v in PRICE_UPDATES.items():
        if sheet[f'A{row}'].value == k:
            sheet[f'B{row}'] = v
wb.save('produce_sales_updated.xlsx')

print(time() - start_time)
