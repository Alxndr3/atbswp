#! /bin/env python3
import sys
import openpyxl


# if len(sys.argv) < 2:
#     print("Type the name of file to invert row with column.")
# else:

# wb = openpyxl.load_workbook(sys.argv[0])
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
wbn = openpyxl.Workbook()
sheetn = wbn.active

item_row = list()
for column_ in range(1, sheet.max_column + 1):
    for row_ in range(1, sheet.max_row + 1):
        sheetn.cell(row=column_, column=row_).value = sheet.cell(row=row_, column=column_).value

wbn.save('new_example.xlsx')
