#! /usr/bin/env python3
# spreadsheet_to_file.py - Converts each column from a spreadsheet into text
# text file
import openpyxl
from openpyxl.utils import get_column_letter
from sys import argv



if len(argv) < 2:
    print("Usage: ./spreadsheet_to_text.py file_to_convert_to_txt.xlsx")
else:
    wb = openpyxl.load_workbook(argv[1])
    sheet = wb.active
    for x in range(1, sheet.max_column + 1):
        with open('%s_column_%s.txt' % (argv[1], get_column_letter(x)), 'w') as tx:
            for y in range(1, sheet.max_row, + 1):
                print(get_column_letter(x))
                tx.write(str(sheet.cell(row=y, column=x).value) + '\n')
                print(str(sheet.cell(row=y, column=x).value))

