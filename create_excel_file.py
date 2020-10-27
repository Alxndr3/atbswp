import openpyxl

# # Createing and saving excel documents.
# wb = openpyxl.Workbook() # Create a blank workbook
# wb.sheetnames # It starts with one sheet.
# sheet = wb.active
# print(sheet.title)
# sheet.title = 'Spam Bacon Eggs Sheet' # Change title.
# print(sheet.title)
# print(wb.sheetnames)
# wb.save('new_file.xlsx') # Save the workbook

# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb.active
# sheet.title = 'Spam Spam Spam'
# wb.save('example_copy.xlsx') # Save the workbook
# print(wb.sheetnames)

# # Creating and removing sheets.
# wb = openpyxl.Workbook()
# wb.create_sheet() # Add a new sheet.

# # Create a new sheet at index 0.
# wb.create_sheet(index=0, title="First Sheet")
# wb.create_sheet(index=2, title="Middle Sheet")
# print(wb.sheetnames)
# # Delete sheet from workbook.
# del wb['Middle Sheet']
# del wb['Sheet1']
# print(wb.sheetnames)

# Writing values to cells.
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hello, World!' # Edit the cell's value.
print(sheet['A1'].value)
wb.save('hello_world.xlsx')
