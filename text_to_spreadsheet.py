import os
import openpyxl


wb = openpyxl.Workbook()
sheet = wb.active
c = sheet.max_column
for y in os.listdir():
    if y.endswith('.txt'):
        r = 1
        with open(y, 'r', encoding="ISO-8859-1") as text_file:
            for x in text_file.readlines():
                print(sheet.max_column)
                sheet.cell(row=r, column=c).value = str(x)
                r += 1
            c += 1

wb.save('textToExcel.xslx')
